import openpyxl
from collections import Counter,defaultdict

path_database = r'C:\Users\burca\OneDrive\Desktop\UNI\ANUL_III\SEM_1\AI\Tema1\Data cat personality and predation Cordonnier et al.xlsx'

workbook = openpyxl.load_workbook(path_database)
sheet = workbook['Data']

#---------------------------------------------------------------
def numara_instante(sheet, column_index):
    rase = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        rasa = row[column_index - 1]
        if rasa is not None:
            rase.append(rasa)

    numar_instante = Counter(rase)

    return numar_instante


column_index = 5

numar_instante = numara_instante(sheet, column_index)

print("Numărul de instanțe pentru fiecare rasă de pisici:")
for rasa, count in numar_instante.items():
    print(f"{rasa}: {count}")
print("------------------------------------------------------------------------------------")

#--------------------------------------------------------------------------
def frecvente_per_rasa(sheet):
    numar_col = sheet.max_column
    frecvente_rasa = defaultdict(lambda: defaultdict(Counter))

    header = [cell.value for cell in sheet[1]] #antet

    for row in sheet.iter_rows(min_row=2, values_only=True):
        rasa = row[4]  #coloana cu rasa

        for col_index in range(3, numar_col):
            if col_index == 5:
                continue  #sarim peste coloana cu  rasa

            atribut = row[col_index - 1]
            if atribut is not None:
                frecvente_rasa[rasa][header[col_index - 1]][atribut] += 1

    return frecvente_rasa

def afiseaza_frecvente_si_procente(sheet):
    frecvente_rasa = frecvente_per_rasa(sheet)

    for rasa, frecvente_per_col in frecvente_rasa.items():
        print(f"\nRasă: {rasa}")

        for col_nume, frecvente in frecvente_per_col.items():
            total_instante_col = sum(frecvente.values())
            print(f"  Coloana {col_nume}:")

            for atribut, numar in frecvente.items():
                procent = (numar / total_instante_col) * 100
                print(f"    Valoarea {atribut}: {numar} instanțe, {procent:.2f}%")

afiseaza_frecvente_si_procente(sheet)
print("------------------------------------------------------------------------------------")

#--------------------------------------------------------------------------------------------

def frecvente_per_fisier(sheet):
    numar_col = sheet.max_column
    frecvente_fisier = {}

    header = [cell.value for cell in sheet[1]]

    for col_index in range(3, numar_col):
        frecvente_fisier[header[col_index - 1]] = Counter()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for col_index in range(3, numar_col):
            atribut = row[col_index - 1]
            if atribut is not None:
                frecvente_fisier[header[col_index - 1]][atribut] += 1

    return frecvente_fisier


def afiseaza_frecvente_fisier(sheet):
    frecvente_fisier = frecvente_per_fisier(sheet)

    for col_nume, frecvente in frecvente_fisier.items():
        total_instante_col = sum(frecvente.values())
        print(f"\nColoana '{col_nume}':")

        for atribut, numar in frecvente.items():
            procent = (numar / total_instante_col) * 100
            print(f"  Valoarea '{atribut}': {numar} instanțe, {procent:.2f}%")


afiseaza_frecvente_fisier(sheet)